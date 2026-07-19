import datetime
import logging
import os
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from payment.models import FixedAsset, DepreciationEntry

logger = logging.getLogger(__name__)

# Column layout (0-indexed) as found in the real depreciation register
# (`ค่าเสื่อมราคา_06-2026.xlsx`, sheet named after the Buddhist year e.g. "2569"):
#   0  ลำดับ                      (item no)
#   1  ทรัพย์สิน                  (asset code, e.g. "2-CAR-19-002")
#   2  ใบสำคัญจ่าย (PV number)    -> legacy_pv_ref
#   3  วันที่จ่าย                 (payment date)
#   4  วันที่เริ่มคำนวณ           (depreciation start date) -> purchase_date
#   5  รายการทรัพย์สิน            (description / name)
#   6  Supplier
#   7  ประเภท                    (sub-type, informational only)
#   8  ราคาทรัพย์สินที่ซื้อมา     (purchase price)
#   9  ค่าเสื่อมราคาสะสมยกมา     (accumulated depreciation brought forward)
#  10  อัตรา (ร้อยละ)             (rate, either 0.2 or 20 form)
#  11  TOTAL                      (annual depreciation total)
#  12  Days
#  13-24  JAN..DEC                (monthly depreciation amounts)
#  25  DEPN Total
#  26  Acc. Depn (as of report date)
#  27  Net Book Value
#  28  TAX INV.#
#  29  Location / user remark
COL_ITEM_NO = 0
COL_ASSET_CODE = 1
COL_PV_REF = 2
COL_PAID_DATE = 3
COL_START_DATE = 4
COL_NAME = 5
COL_SUPPLIER = 6
COL_SUBTYPE = 7
COL_PRICE = 8
COL_ACC_DEPN_BROUGHT_FORWARD = 9
COL_RATE = 10
COL_MONTHS_START = 13
COL_MONTHS_END = 24  # inclusive
COL_TAX_INV = 28

MONTH_NAMES = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

CATEGORY_KEYWORDS = [
    ('ยานพาหนะ', 'VEHICLE'),
    ('โปรแกรมคอมพิวเตอร์', 'SOFTWARE'),
    ('ซอฟต์แวร์', 'SOFTWARE'),
    ('คอมพิวเตอร์', 'COMPUTER'),
    ('ตกแต่งสำนักงาน', 'OFFICE_DECOR'),
    ('เครื่องตกแต่ง', 'FURNITURE'),
    ('เครื่องมือเครื่องใช้', 'FURNITURE'),
]


def _map_category(label):
    if not label:
        return 'OTHER'
    for keyword, category in CATEGORY_KEYWORDS:
        if keyword in label:
            return category
    return 'OTHER'


def _to_decimal(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def _to_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


class Command(BaseCommand):
    help = (
        "Import fixed assets and their monthly depreciation entries from the legacy Excel "
        "depreciation register (one sheet with section headers per asset category)."
    )

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Path to the xlsx file to import.')
        parser.add_argument(
            '--sheet', default=None,
            help='Sheet name to import (defaults to the first numeric-named sheet, e.g. "2569").',
        )
        parser.add_argument(
            '--year', type=int, default=None,
            help='Calendar (AD) year the JAN-DEC columns represent. Defaults to sheet name interpreted '
                 'as a Buddhist Era year (sheet_name - 543).',
        )

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            raise CommandError(f"File not found: {file_path}")

        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        sheet_name = options.get('sheet')
        if not sheet_name:
            sheet_name = self._guess_sheet_name(wb)
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

        year = options.get('year')
        if year is None:
            year = self._guess_year(sheet_name)
        if year is None:
            raise CommandError(
                "Could not determine calendar year from sheet name; pass --year explicitly."
            )

        ws = wb[sheet_name]

        assets_created = 0
        assets_updated = 0
        entries_created = 0
        current_category = 'OTHER'

        for row in ws.iter_rows(values_only=True):
            if row is None or all(v is None for v in row):
                continue

            item_no = row[COL_ITEM_NO] if len(row) > COL_ITEM_NO else None
            asset_code = row[COL_ASSET_CODE] if len(row) > COL_ASSET_CODE else None
            price = row[COL_PRICE] if len(row) > COL_PRICE else None

            # Category header row: no item_no, asset_code column holds a small int
            # (the category number), and no purchase price.
            if item_no is None and isinstance(asset_code, (int, float)) and price is None:
                label = row[COL_NAME] if len(row) > COL_NAME else None
                current_category = _map_category(label)
                continue

            # Data row: must have an asset code (string) and a numeric purchase price.
            if not isinstance(asset_code, str) or not isinstance(price, (int, float)):
                continue

            purchase_price = _to_decimal(price)
            if purchase_price is None:
                continue

            purchase_date = (
                _to_date(row[COL_START_DATE]) if len(row) > COL_START_DATE else None
            ) or (
                _to_date(row[COL_PAID_DATE]) if len(row) > COL_PAID_DATE else None
            ) or datetime.date(year, 1, 1)

            rate_raw = _to_decimal(row[COL_RATE]) if len(row) > COL_RATE else None
            if rate_raw is None:
                rate_percent = Decimal('0')
            elif rate_raw <= 1:
                rate_percent = (rate_raw * 100).quantize(Decimal('0.01'))
            else:
                rate_percent = rate_raw

            pv_ref = row[COL_PV_REF] if len(row) > COL_PV_REF else None
            tax_inv = row[COL_TAX_INV] if len(row) > COL_TAX_INV else None
            name = row[COL_NAME] if len(row) > COL_NAME and row[COL_NAME] else asset_code

            with transaction.atomic():
                asset, created = FixedAsset.objects.update_or_create(
                    asset_code=asset_code,
                    defaults={
                        'name': str(name)[:255],
                        'category': current_category,
                        'legacy_pv_ref': str(pv_ref)[:50] if pv_ref else '',
                        'tax_invoice_no': str(tax_inv)[:100] if tax_inv else '',
                        'purchase_price': purchase_price,
                        'purchase_date': purchase_date,
                        'depreciation_rate_percent': rate_percent,
                    },
                )
                if created:
                    assets_created += 1
                else:
                    assets_updated += 1

                accumulated = _to_decimal(row[COL_ACC_DEPN_BROUGHT_FORWARD]) if len(row) > COL_ACC_DEPN_BROUGHT_FORWARD else None
                accumulated = accumulated or Decimal('0')

                for month_idx, col in enumerate(range(COL_MONTHS_START, COL_MONTHS_END + 1)):
                    if col >= len(row):
                        continue
                    monthly_amount = _to_decimal(row[col])
                    if not monthly_amount:
                        continue
                    accumulated += monthly_amount
                    period = f"{year}-{month_idx + 1:02d}"
                    _, entry_created = DepreciationEntry.objects.update_or_create(
                        asset=asset,
                        period=period,
                        defaults={
                            'amount': monthly_amount,
                            'accumulated_amount': accumulated,
                        },
                    )
                    if entry_created:
                        entries_created += 1

        self.stdout.write(self.style.SUCCESS(
            f"Import complete. Assets created: {assets_created}, updated: {assets_updated}, "
            f"depreciation entries created: {entries_created}."
        ))

    def _guess_sheet_name(self, wb):
        for name in wb.sheetnames:
            if name.isdigit():
                return name
        return wb.sheetnames[0]

    def _guess_year(self, sheet_name):
        if sheet_name.isdigit():
            be_year = int(sheet_name)
            if be_year > 2400:
                return be_year - 543
            return be_year
        return None
