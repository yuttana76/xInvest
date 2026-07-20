from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from django.utils.formats import date_format
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import WorkflowConfig, WorkflowStep, Request, ApprovalLog, RequestFile, RequestSubject
from .tasks import dispatch_workflow_action_required_email


class WorkflowStepInline(admin.TabularInline):
    model = WorkflowStep
    extra = 1


class RequestSubjectInline(admin.TabularInline):
    model = RequestSubject
    extra = 1
    fields = ('order', 'name', 'code', 'is_active')


class ApprovalLogInline(admin.TabularInline):
    model = ApprovalLog
    readonly_fields = ('approver', 'step_number', 'step_name', 'action', 'comment', 'created_at')
    extra = 0
    can_delete = False


class RequestFileInline(admin.TabularInline):
    model = RequestFile
    extra = 1


@admin.register(RequestSubject)
class RequestSubjectAdmin(admin.ModelAdmin):
    list_display = ('order', 'workflow', 'name', 'code', 'is_active')
    list_display_links = ('name',)
    list_editable = ('order', 'is_active')
    list_filter = ('workflow', 'is_active')
    search_fields = ('name', 'code', 'workflow__name')


@admin.register(WorkflowConfig)
class WorkflowConfigAdmin(admin.ModelAdmin):
    list_display = ('name', 'category')
    inlines = [WorkflowStepInline, RequestSubjectInline]


@admin.register(Request)
class RequestAdmin(admin.ModelAdmin):
    list_display = ('req_code', 'title', 'workflow', 'creator', 'status', 'priorify', 'expectDate', 'auditFlag', 'get_current_step', 'created_at')
    list_filter = ('status', 'workflow', 'priorify', 'auditFlag')
    search_fields = ('req_code', 'title', 'creator__username', 'creator__email')
    inlines = [ApprovalLogInline, RequestFileInline]
    readonly_fields = ('creator', 'created_at', 'updated_at', 'completed_at')
    filter_horizontal = ('reqSubject',)
    actions = ['export_as_excel', 'resend_approval_email']

    def get_current_step(self, obj):
        step = obj.get_current_step_info()
        if step:
            return f"Step {obj.current_step_number}: {step.step_name}"
        return f"Step {obj.current_step_number}"
    get_current_step.short_description = 'Current Step'

    def resend_approval_email(self, request, queryset):
        for req in queryset:
            current_step = req.get_current_step_info()
            if not current_step:
                self.message_user(request, f"Request {req.req_code or req.id} has no current step to notify.", level='warning')
                continue

            if current_step.is_department_manager:
                manager = None
                if hasattr(req.creator, 'profile') and req.creator.profile.department:
                    manager = req.creator.profile.department.manager
                if manager:
                    dispatch_workflow_action_required_email(req.id, user_id=manager.id)
                else:
                    self.message_user(request, f"No department manager found for request {req.req_code or req.id}.", level='warning')
            elif current_step.required_group:
                dispatch_workflow_action_required_email(req.id, group_id=current_step.required_group.id)
            else:
                self.message_user(request, f"No approver group configured for request {req.req_code or req.id}.", level='warning')

        self.message_user(request, f"Approval reminder emails queued for {queryset.count()} request(s).", level='success')

    resend_approval_email.short_description = 'Resend approval email'

    def export_as_excel(self, request, queryset):
        headers = [
            'Req code', 'Title', 'Workflow', 'Creator', 'Department', 'Status',
            'Priority', 'Expected Completion Date', 'Audit Flag', 'Current Step',
            'Created at', 'Approval logs',
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Requests'
        ws.append(headers)

        queryset = queryset.select_related('workflow', 'creator').prefetch_related('logs__approver')
        for req in queryset:
            step_info = req.get_current_step_info()
            current_step = f"Step {req.current_step_number}: {step_info.step_name}" if step_info else f"Step {req.current_step_number}"

            approval_logs = ''.join(
                f"{log.step_number}:{log.approver if log.approver else '-'}:{log.get_action_display()};"
                for log in req.logs.order_by('step_number')
            )

            created_at_local = timezone.localtime(req.created_at) if req.created_at else None

            ws.append([
                req.req_code,
                req.title,
                req.workflow.name if req.workflow else '',
                str(req.creator) if req.creator else '',
                req.create_department or '',
                req.get_status_display(),
                req.get_priorify_display(),
                req.expectDate.strftime('%d-%b-%y') if req.expectDate else '',
                'TRUE' if req.auditFlag else 'FALSE',
                current_step,
                date_format(created_at_local, 'DATETIME_FORMAT') if created_at_local else '',
                approval_logs,
            ])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 24

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="workflow_requests.xlsx"'
        wb.save(response)
        return response
    export_as_excel.short_description = 'Export selected requests to Excel'


@admin.register(ApprovalLog)
class ApprovalLogAdmin(admin.ModelAdmin):
    list_display = ('request', 'approver', 'step_name', 'action', 'created_at')
    list_filter = ('action', 'created_at')
    search_fields = ('request__title', 'approver__username', 'comment')


@admin.register(WorkflowStep)
class WorkflowStepAdmin(admin.ModelAdmin):
    list_display = ('step_name', 'workflow', 'step_number', 'required_group', 'is_department_manager')
    list_filter = ('workflow', 'required_group', 'is_department_manager')
    search_fields = ('step_name', 'workflow__name')


@admin.register(RequestFile)
class RequestFileAdmin(admin.ModelAdmin):
    list_display = ('request', 'file', 'file_name', 'uploaded_at')
    list_filter = ('request', 'uploaded_at')
    search_fields = ('request__title', 'file_name')

